import os
import SimpleITK as sitk
import pandas as pd
from openpyxl import load_workbook

def update_excel_with_mha(folder_path, excel_path,sheet_name):


    # wczytanie notatnika excel
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # pobieranie wszystkich plików do sprawdzenia
    files = [f for f in os.listdir(folder_path)]
    
    print("Pobrano pliki oraz otwarto excela")


    row = 2
    for idx, file in enumerate(files):


        file_path = os.path.join(folder_path,file)
        if not os.path.exists(file_path):
            print("Niez znaleziono pliku: " + file_path)
            continue
        #print(str(idx) + str(file))
        print("Operowanie na pliku " + file)
        _, rozszerzenie = os.path.splitext(file_path)
        image = sitk.ReadImage(file_path)
        size = image.GetSize()
        spacing = image.GetSpacing()
        dtype = image.GetPixelIDTypeAsString()
        channels = image.GetNumberOfComponentsPerPixel()
        array = sitk.GetArrayFromImage(image)
        min_val, max_val = array.min(), array.max()
        ws[f"A{row}"] = file
        ws[f"B{row}"] = rozszerzenie
        ws[f"C{row}"] = f"{size[0]}x{size[1]}x{size[2]}"
        ws[f"D{row}"] = f"{spacing[0]:.3f}x{spacing[1]:.3f}x{spacing[2]:.3f}"
        ws[f"E{row}"] = channels
        ws[f"F{row}"] = dtype
        ws[f"G{row}"] = min_val
        ws[f"H{row}"] = max_val

        row = row + 1

    wb.save(excel_path)
    print(f"Zaktualizowano plik excel: {excel_path}")

    
    # Tworzenie DataFrame i zapis do Excela


data_folder = "E:\\magisterka\\projekt\\CBCT_bone_segmentation\\Data\\ToothFairy2\\imagesTr"
excel_file = "C:\\Users\\karno\\Desktop\\dane.xlsx"
sheet_name = "ToothFairy2"

update_excel_with_mha(data_folder,excel_file,sheet_name)

#print(f"Plik zapisany: {output_excel}")
