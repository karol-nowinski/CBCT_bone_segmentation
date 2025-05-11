import torch
import torchio as tio
import numpy as np
from pathlib import Path
import config
import SimpleITK as sitk
import pandas as pd
from openpyxl import load_workbook




def compute_metrics_per_label(ground_truth, prediction, labels):
    metrics = {}

    for label in labels:
        gt_label = sitk.GetArrayFromImage(ground_truth == label).astype(np.uint8)
        pred_label = sitk.GetArrayFromImage(prediction == label).astype(np.uint8)

        intersection = np.logical_and(gt_label, pred_label).sum()
        union = np.logical_or(gt_label, pred_label).sum()
        gt_sum = gt_label.sum()
        pred_sum = pred_label.sum()
        total = gt_label.size

        dice = 2 * intersection / (gt_sum + pred_sum) if (gt_sum + pred_sum) > 0 else 1.0
        iou = intersection / union if union > 0 else 1.0
        recall = intersection / gt_sum if gt_sum > 0 else 1.0
        accuracy = (gt_label == pred_label).sum() / total

        metrics[label] = {
            "Dice": dice,
            "IoU": iou,
            "Recall": recall,
            "Accuracy": accuracy
        }

    return metrics


def prepare_paths(test_folder,prediction_folder,extension):

    test_dir = Path(test_folder)
    pred_dir = Path(prediction_folder)

    pairs = []

    for file_path in test_dir.glob('*'+extension):
        splited_name = file_path.name.split('.')
        searched_name = splited_name[0] + "_0000_prediction" + extension
        prediction_path = pred_dir / searched_name
        pairs.append((file_path,prediction_path))

    return pairs

def save_metrics_to_excel(results, excel_path):

    rows = []

    for file_info in results:
        gt_path = file_info["ground_truth_path"]
        pred_path = file_info["prediction_path"]
        metrics = file_info["metrics"]

        for label, values in metrics.items():
            row = {
                "Ground Truth Path": gt_path,
                "Prediction Path": pred_path,
                "Label": label
            }
            row.update(values)
            rows.append(row)

    df = pd.DataFrame(rows)

    try:
        # Jeśli plik istnieje — dopisz do niego
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = load_workbook(excel_path)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            start_row = writer.sheets['Arkusz1'].max_row
            df.to_excel(writer, index=False, header=False, startrow=start_row)
    except FileNotFoundError:
        # Jeśli plik nie istnieje — zapisz jako nowy
        df.to_excel(excel_path, index=False)



if __name__ == "__main__":
    print("--- Uruchomienie skryptu liczenia metryk testowych ---")

    test_folder = "Data\\CleanToothFairy2\\labelsTeethAll\\test"
    prediction_folder = "Results\\CleanToothFairy_MergedTeeth_Unet3DPP96" 


    pairs = prepare_paths(test_folder,prediction_folder,config.FILE_FORMAT)

    #print(pairs)

    results_per_file = []
    for gr, pred in pairs:
        print(f"Operowanie na pliku: {gr}")

        gt_image = sitk.ReadImage(gr)
        pred_image = sitk.ReadImage(pred)

        metrics = compute_metrics_per_label(gt_image,pred_image,list(range(1, 9)))
        print(metrics)
        results_per_file.append({
            "metrics": metrics,
            "ground_truth_path": gr,
            "prediction_path": pred
        })


    save_metrics_to_excel(results_per_file,"toothfairy2_mergedTeeth_UnetPP3D.xlsx")





