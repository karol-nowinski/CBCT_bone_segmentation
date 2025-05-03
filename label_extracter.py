import os
import SimpleITK as sitk
import numpy as np
from openpyxl import load_workbook

# extracter operujący na plikach zaiwrających maski segmentacyjne

def analyze_mask(mask_array):
    unique_values, counts = np.unique(mask_array, return_counts=True)
    class_list = unique_values.tolist()
    voxel_counts = counts.tolist()
    return len(class_list), class_list, voxel_counts


def update_excel_with_mask(labels_path, excel_file, sheet_name):

    # wczytanie excela
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # pobieranie wszytskich plików
    files = [f for f in os.listdir(labels_path)]
    print("Pobrano pliki oraz otawrto excela")

    row = 2

    for idx, file in enumerate(files):


        file_path = os.path.join(label_folder,file)
        if not os.path.exists(file_path):
            print("Niez znaleziono pliku: " + file_path)
            continue     
        
        print("Operowanie na pliku " + file)


        mask = sitk.ReadImage(file_path)
        mask_array = sitk.GetArrayFromImage(mask)

        size = mask.GetSize()
        spacing = mask.GetSpacing()


        num_classes, class_list, voxel_counts = analyze_mask(mask_array)

        # wypełnianie wiersza excela
        ws[f"A{row}"] = file
        ws[f"B{row}"] = f"{size[0]}x{size[1]}x{size[2]}"
        ws[f"C{row}"] = f"{spacing[0]:.3f}x{spacing[1]:.3f}x{spacing[2]:.3f}"
        ws[f"D{row}"] = num_classes
        ws[f"E{row}"] = ", ".join(map(str, class_list))
        ws[f"F{row}"] = ", ".join(map(str, voxel_counts))
        print(voxel_counts)

        row = row + 1

    wb.save(excel_file)
    print(f"Zaktualizowano plik excel: {excel_file}")


label_folder =  "E:\\magisterka\\projekt\\CBCT_bone_segmentation\\Data\\ToothFairy2\\labelsTr"
excel_file = "C:\\Users\\karno\\Desktop\\dane.xlsx"
sheet_name = "ToothFairyLabel"

update_excel_with_mask(label_folder,excel_file,sheet_name)